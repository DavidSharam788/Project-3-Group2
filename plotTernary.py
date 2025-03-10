import mpltern
import matplotlib.pyplot as plt
import openpyxl
import numpy as np


path = "sampleBA0.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
dat_len = sheet_obj.max_row
gen = np.zeros(dat_len)
con = np.zeros(dat_len)
pas = np.zeros(dat_len)
kc = np.zeros(dat_len)
for i in range(dat_len):
    gen[i] = sheet_obj.cell(row = i + 1, column = 1).value
    con[i] = sheet_obj.cell(row = i + 1, column = 2).value
    pas[i] = sheet_obj.cell(row = i + 1, column = 3).value
    kc[i] = sheet_obj.cell(row = i + 1, column = 4).value

fig = plt.figure()
ax = plt.subplot(projection="ternary", ternary_sum = 20)
ax.set_tlabel("generator")
ax.set_llabel("consumer")
ax.set_rlabel("passive")
ax.grid()

tick_values = [0, 5, 10, 15, 20]
ax.taxis.set_ticks(tick_values)
ax.laxis.set_ticks(tick_values)
ax.raxis.set_ticks(tick_values)

ax.taxis.set_label_position('tick2')
ax.laxis.set_label_position('tick2')
ax.raxis.set_label_position('tick2')
plt.title(r"$\kappa_c$ for Barabasi-Albert Lattice Networks (n = 20)")



pc = ax.tricontourf(pas, gen, con, kc)

cax = ax.inset_axes([1.05, 0.1, 0.05, 0.9], transform=ax.transAxes)
colorbar = fig.colorbar(pc, cax=cax)
colorbar.set_label(r"$\kappa_c$", rotation=270, va="baseline")

plt.show()