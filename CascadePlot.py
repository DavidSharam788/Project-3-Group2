import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = 'CascadeData.xlsx')
sheet_obj = wb.active
data_width = 10
dat_len = sheet_obj.max_row
acastars = []
Data1 = np.zeros(data_width)
Data2 = np.zeros(data_width)
Data3 = np.zeros(data_width)
Data4 = np.zeros(data_width)

for i in range(dat_len):
    for j in range(data_width):
        if(i == 0):
            acastars.append(sheet_obj.cell(row = i + 1, column = j + 1).value)
        elif((i-1)%4 == 0):
            Data1[j] += float(sheet_obj.cell(row = i + 1, column = j + 1).value)/10
        elif((i-1)%4 == 1):
            Data2[j] += int(sheet_obj.cell(row = i + 1, column = j + 1).value)/590
        elif((i-1)%4 == 2):
            Data3[j] += int(sheet_obj.cell(row = i + 1, column = j + 1).value)/590
        elif((i-1)%4 == 3):
            Data4[j] += float(sheet_obj.cell(row = i + 1, column = j + 1).value)/10
#plt.plot(acastars,Data1,label = 'S')
#plt.plot(acastars,Data2,label = 'Overloads')
#plt.plot(acastars,Data3,label = 'Desyncs')
plt.plot(acastars,Data4)
plt.legend()
plt.ylabel("Cascade duration")
plt.xlabel(r"$\alpha_c$")
plt.show()